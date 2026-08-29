"""Excel export with styling."""
from __future__ import annotations
import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporterService:
    @staticmethod
    def export_results_to_excel(df_results: pd.DataFrame, risk_threshold: float = 5.0) -> bytes:
        output = io.BytesIO()
        margin_col = None
        for c in df_results.columns:
            cl = str(c).lower()
            if "маржинальность" in cl:
                margin_col = c
                break
        if margin_col is None:
            for c in df_results.columns:
                if "рентабельность" in str(c).lower():
                    margin_col = c
                    break

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_results.to_excel(writer, index=False, sheet_name="Все")
            if margin_col is not None:
                risk = df_results[df_results[margin_col] < risk_threshold]
                top = df_results.sort_values(by=margin_col, ascending=False).head(50)
            else:
                profit_col = "Чистая прибыль, ₽" if "Чистая прибыль, ₽" in df_results.columns else None
                if profit_col:
                    risk = df_results[df_results[profit_col] < 0]
                    top = df_results.sort_values(by=profit_col, ascending=False).head(50)
                else:
                    risk = df_results.iloc[0:0]
                    top = df_results.head(50)
            risk.to_excel(writer, index=False, sheet_name="Риск")
            top.to_excel(writer, index=False, sheet_name="Топ")

        output.seek(0)
        wb = openpyxl.load_workbook(output)

        # Лист со справкой по терминам
        if "Справка" in wb.sheetnames:
            del wb["Справка"]
        ws_help = wb.create_sheet("Справка", 0)
        help_rows = [
            ("Термин", "Что значит в Маржинаторе"),
            ("Выручка, ₽", "Цена продажи (с учётом модели расчёта)"),
            ("Себестоимость, ₽", "Закуп (после курса валюты, если задавали)"),
            ("Переменные расходы, ₽", "Закуп + комиссия + логистика + упаковка и др."),
            ("Маржа, ₽", "Выручка − переменные (до налога)"),
            ("Маржинальность %", "Маржа ÷ выручка × 100%. Доля выручки после переменных"),
            ("Наценка %", "Маржа ÷ переменные × 100%. Не путать с маржинальностью"),
            ("Чистая прибыль, ₽", "Маржа − налог (оценка «в карман»)"),
            ("Рентабельность чистая %", "Чистая прибыль ÷ выручка × 100%"),
            ("ROI %", "Чистая прибыль ÷ себестоимость × 100%"),
            ("Цвет строки", f"🟢 ≥ 20% · 🟡 {risk_threshold:g}–20% · 🔴 < {risk_threshold:g}%"),
            ("Лист «Риск»", f"Позиции с маржинальностью < {risk_threshold:g}%"),
            ("Лист «Топ»", "Лучшие позиции по маржинальности"),
        ]
        # Title row
        ws_help.insert_rows(1)
        title = ws_help.cell(row=1, column=1, value="Marginator — отчёт юнит-экономики")
        title.font = Font(name="Calibri", size=14, bold=True, color="064E3B")
        ws_help.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        ws_help.cell(row=2, column=1, value="Термин")
        ws_help.cell(row=2, column=2, value="Что значит")
        for r, (a, b) in enumerate(help_rows[1:], 3):  # skip old header pair
            ws_help.cell(row=r, column=1, value=a)
            ws_help.cell(row=r, column=2, value=b)
        # header style
        for col in (1, 2):
            c = ws_help.cell(row=2, column=col)
            c.fill = PatternFill(start_color="064E3B", end_color="064E3B", fill_type="solid")
            c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        ws_help.column_dimensions["A"].width = 28
        ws_help.column_dimensions["B"].width = 72
        ws_help.freeze_panes = "A3"
        ws_help.row_dimensions[1].height = 22


        header_fill = PatternFill(start_color="064E3B", end_color="064E3B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
        green_font = Font(name="Calibri", color="065F46", bold=True)
        yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        yellow_font = Font(name="Calibri", color="92400E")
        red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        red_font = Font(name="Calibri", color="991B1B", bold=True)
        zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="E5E7EB"),
            right=Side(style="thin", color="E5E7EB"),
            top=Side(style="thin", color="E5E7EB"),
            bottom=Side(style="thin", color="E5E7EB"),
        )

        def _norm(val) -> str:
            return str(val or "").strip().lower().replace("ё", "е")

        def style_sheet(ws, risk_sheet: bool = False):
            if ws.max_row < 1 or ws.max_column < 1:
                return
            ws.row_dimensions[1].height = 30
            margin_cols, profit_cols, roi_cols, highlight = [], [], [], set()
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                name = _norm(cell.value)
                if any(x in name for x in ("маржинальность", "маржин", "маржа, ₽", "маржа ₽", "margin", "рентаб", "наценка")):
                    margin_cols.append(col_num)
                    highlight.add(col_num)
                if any(x in name for x in ("прибыл", "profit", "net")):
                    profit_cols.append(col_num)
                    highlight.add(col_num)
                if "roi" in name:
                    roi_cols.append(col_num)
                    highlight.add(col_num)

            for row_num in range(2, ws.max_row + 1):
                ws.row_dimensions[row_num].height = 20
                margin_val = None
                for mc in margin_cols:
                    v = ws.cell(row=row_num, column=mc).value
                    if isinstance(v, (int, float)):
                        margin_val = float(v)
                        break
                if margin_val is None:
                    for pc in profit_cols:
                        v = ws.cell(row=row_num, column=pc).value
                        if isinstance(v, (int, float)):
                            margin_val = 20.0 if float(v) > 0 else -1.0
                            break

                if risk_sheet:
                    row_fill, row_font = red_fill, red_font
                elif margin_val is None:
                    row_fill = zebra if row_num % 2 == 0 else None
                    row_font = None
                elif margin_val >= 20.0:
                    row_fill, row_font = green_fill, green_font
                elif margin_val >= risk_threshold:
                    row_fill, row_font = yellow_fill, yellow_font
                else:
                    row_fill, row_font = red_fill, red_font

                for col_num in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")
                    if col_num in highlight and row_fill is not None:
                        cell.fill = row_fill
                        if row_font:
                            cell.font = row_font
                    elif row_num % 2 == 0 and col_num not in highlight:
                        cell.fill = zebra
                    if isinstance(cell.value, (int, float)):
                        name = _norm(ws.cell(row=1, column=col_num).value)
                        if any(x in name for x in ("%", "маржин", "roi", "наценка", "рентаб")):
                            cell.number_format = "0.00"
                        elif any(x in name for x in ("₽", "руб", "цена", "прибыл", "себестоим", "выруч", "маржа", "перемен")):
                            cell.number_format = "#,##0.00"

            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, min(len(str(cell.value)), 48))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            ws.freeze_panes = "A2"
            if ws.max_row >= 1:
                ws.auto_filter.ref = ws.dimensions

        for name in wb.sheetnames:
            if name == "Справка":
                continue
            style_sheet(wb[name], risk_sheet=(name == "Риск"))

        final = io.BytesIO()
        wb.save(final)
        return final.getvalue()

    @staticmethod
    def export_buy_list(df_results: pd.DataFrame, min_roi: float = 30.0) -> bytes:
        df = df_results.copy()
        if "ROI %" in df.columns:
            filtered = df[df["ROI %"] >= min_roi]
        elif "Маржинальность %" in df.columns:
            filtered = df[df["Маржинальность %"] >= 5.0]
        else:
            filtered = df
        if filtered.empty:
            filtered = df.head(0)
        return ExcelExporterService.export_results_to_excel(filtered)
