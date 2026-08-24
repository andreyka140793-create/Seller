import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelExporterService:
    @staticmethod
    def export_results_to_excel(df_results: pd.DataFrame) -> bytes:
        """
        Генерирует стилизованный .xlsx файл с подсвеченными показателями маржи.
        Возвращает бинарные данные файла для отправки в Telegram.
        """
        output = io.BytesIO()
        
        # Сохраняем первичный DataFrame в буфер
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_results.to_excel(writer, index=False, sheet_name='Маржинатор_Отчет')
            
        output.seek(0)
        wb = openpyxl.load_workbook(output)
        ws = wb['Маржинатор_Отчет']

        # --- Цветовая палитра и шрифты ---
        header_fill = PatternFill(start_color="064E3B", end_color="064E3B", fill_type="solid")  # Изумрудный
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")   # Высокая маржа
        green_font = Font(name="Calibri", color="065F46", bold=True)
        
        yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Средняя маржа
        yellow_font = Font(name="Calibri", color="92400E")
        
        red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")     # Отрицательная / Низкая
        red_font = Font(name="Calibri", color="991B1B", bold=True)
        
        thin_border = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB')
        )

        # --- Стилизация заголовков ---
        ws.row_dimensions[1].height = 28
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Поиск ключевых колонок для форматирования
        margin_col_idx = None
        for col_num in range(1, ws.max_column + 1):
            col_name = str(ws.cell(row=1, column=col_num).value).lower()
            if 'маржа' in col_name or 'margin' in col_name:
                margin_col_idx = col_num

        # --- Стилизация данных и цветовая подсветка ---
        for row_num in range(2, ws.max_row + 1):
            ws.row_dimensions[row_num].height = 20
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

                # Подсветка и форматирование колонки «Маржинальность %»
                if col_num == margin_col_idx and isinstance(cell.value, (int, float)):
                    val = float(cell.value)
                    cell.number_format = '0.00"%"'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    
                    if val >= 20.0:
                        cell.fill = green_fill
                        cell.font = green_font
                    elif val >= 5.0:
                        cell.fill = yellow_fill
                        cell.font = yellow_font
                    else:
                        cell.fill = red_fill
                        cell.font = red_font

        # --- Автоподбор ширины столбцов ---
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 5, 14)

        # Сохранение и возврат байт
        final_output = io.BytesIO()
        wb.save(final_output)
        return final_output.getvalue()
